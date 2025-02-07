import pandas as pd
import logging
from typing import Optional, Dict, Tuple
import os
import re
import openpyxl
import warnings

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Add these helper functions after imports
def get_column_letter(column_number: int) -> str:
    """Convert numeric column index to Excel column letter (e.g., 0 -> 'A', 1 -> 'B', 26 -> 'AA')"""
    result = ""
    while (column_number >= 0):
        column_number, remainder = divmod(column_number, 26)
        result = chr(65 + remainder) + result
        column_number -= 1
    return result

def get_excel_coordinates(row_idx: int, col_idx: int) -> str:
    """Convert numeric row and column indices to Excel-style coordinates (e.g., 0,0 -> 'A1')"""
    return f"{get_column_letter(col_idx)}{row_idx + 1}"

# Set up logging
def setup_logging(log_file: str):
    # Create a logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # Create a file handler to log messages to a file
    file_handler = logging.FileHandler(log_file)
    file_handler.setLevel(logging.INFO)

    # Create a console handler to log messages to the terminal
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)

    # Define a formatter and set it for both handlers
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    # Add both handlers to the logger
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

# Function to extract the community name from the filename
def extract_community_name(file_name: str) -> str:
    # Use a regular expression to extract the community name
    match = re.search(r'\d{4}-\d{2}-(.*?)-reporting-consumptions\.xlsx', file_name)
    if match:
        community_name = match.group(1).replace("-", " ")  # Replace underscores with spaces if present
        return community_name  # Return the modified community name
    return "Unknown"  # Return "Unknown" if no match is found

# Find the last non-empty row in a specific column
def find_last_non_empty_row(df: pd.DataFrame, column: str) -> Optional[int]:
    non_empty_rows = df[column].dropna()
    return non_empty_rows.index[-1] if not non_empty_rows.empty else None

# Modified function to find the first column containing the search text in the correct area
def find_text_cell(df: pd.DataFrame, search_text: str) -> Optional[str]:
    df_subset = df.head(20)
    matching_cols = []
    
    for col in df_subset.columns:
        matches = df_subset[col].astype(str).str.strip() == search_text.strip()
        if matches.any():
            match_row = df_subset.index[matches][0]
            col_idx = int(col.split('_')[1])  # Extract number from 'Column_X'
            matching_cols.append(col)
    
    return matching_cols[0] if matching_cols else None

# Load the result.xlsx file
def load_community_sums(file_path: str) -> pd.DataFrame:
    return pd.read_excel(file_path, engine='openpyxl')

# Process an individual file
def find_value_row(df: pd.DataFrame, column: str, header_row: int) -> Optional[int]:
    """Find the last row with a non-zero numeric value"""
    last_value_row = None
    for idx in reversed(range(len(df))):  # Start from bottom
        value = df.at[idx, column]
        try:
            if pd.notna(value):
                numeric_value = float(str(value).replace(',', ''))
                if numeric_value != 0:
                    # Log surrounding cells for verification
                    col_idx = int(column.split('_')[1])
                    excel_coord = get_excel_coordinates(idx, col_idx)
                    logging.info(f"Found candidate row at {excel_coord} with value {value}")
                    
                    # Show surrounding cells
                    context_start = max(0, idx - 2)
                    context_end = min(len(df), idx + 3)
                    context_df = df.iloc[context_start:context_end, max(0, col_idx-2):col_idx+3]
                    logging.info(f"Context around value:\n{context_df}")
                    
                    last_value_row = idx
                    break
        except (ValueError, TypeError):
            continue
    return last_value_row

def get_formula_sum(df: pd.DataFrame, column: str, start_row: int, end_row: int) -> Optional[int]:
    """Calculate sum manually for a column range"""
    try:
        values = df.iloc[(start_row-1):(end_row), df.columns.get_loc(column)]
        values_list = []
        for idx, val in enumerate(values, start=start_row):
            if pd.notna(val) and str(val).strip():
                try:
                    num_val = float(str(val).replace(',', '').strip())
                    values_list.append(num_val)
                except (ValueError, TypeError):
                    values_list.append(0)
            else:
                values_list.append(0)
        
        return int(sum(values_list))
    except Exception as e:
        logging.error(f"Error calculating sum: {e}")
        return None

def get_countries_sum(values_df: pd.DataFrame, start_col_idx: int, start_row: int, end_row: int) -> int:
    """Calculate sum of all country columns for the given row range"""
    total = 0
    col_idx = start_col_idx
    
    while True:
        col_name = f'Column_{col_idx}'
        if col_idx >= len(values_df.columns):
            break
            
        try:
            values = values_df.iloc[(start_row-1):end_row, col_idx]
            col_total = 0
            for idx, val in enumerate(values, start=start_row):
                if pd.notna(val) and str(val).strip():
                    try:
                        num_val = float(str(val).replace(',', '').strip())
                        col_total += num_val
                    except (ValueError, TypeError):
                        continue
            
            if col_total > 0:
                total += col_total
            else:
                break
                
        except Exception as e:
            logging.warning(f"Error processing column {col_idx}: {e}")
            break
            
        col_idx += 1
    
    return int(total)

def get_row_sum(sheet, row: int, start_col: str) -> int:
    """Sum all non-empty cells in a row starting from given column"""
    total = 0
    col_idx = openpyxl.utils.column_index_from_string(start_col)
    
    while True:
        col_letter = get_column_letter(col_idx - 1)  # -1 because our get_column_letter is 0-based
        cell = sheet[f"{col_letter}{row}"]
        value = cell.value
        
        if value is None or str(value).strip() == '':
            break
            
        try:
            num_val = float(str(value).replace(',', '').strip())
            total += num_val
        except (ValueError, TypeError):
            pass
        
        col_idx += 1
    
    return int(total)

def process_file(file_path: str, search_text1: str, search_text2: str) -> Tuple[str, Optional[int], Optional[int]]:
    try:
        file_name = os.path.basename(file_path)
        community_name = extract_community_name(file_name)

        # Load both versions of the workbook silently
        wb_formulas = openpyxl.load_workbook(filename=file_path, data_only=False)
        wb_values = openpyxl.load_workbook(filename=file_path, data_only=True)
        
        sheet_formulas = wb_formulas["Messaging details"]
        sheet_values = wb_values["Messaging details"]
        
        # Convert formula sheet to DataFrame for searching headers
        data = []
        for row in sheet_formulas.rows:
            data.append([cell.value for cell in row])
        df = pd.DataFrame(data)
        df.columns = [f'Column_{i}' for i in range(df.shape[1])]

        # Find headers
        column_name1 = find_text_cell(df, search_text1)
        column_name2 = find_text_cell(df, search_text2)
        
        if (column_name1 is None) or (column_name2 is None):
            logging.error(f"Could not find headers for {community_name}")
            return community_name, None, None

        # Get the column letters
        col1_letter = get_column_letter(int(column_name1.split('_')[1]))
        col2_letter = get_column_letter(int(column_name2.split('_')[1]))

        # Find formula cells to get their location
        formula_cell1 = None
        formula_cell2 = None
        for row in range(1, sheet_formulas.max_row + 1):
            cell1 = sheet_formulas[f"{col1_letter}{row}"]
            cell2 = sheet_formulas[f"{col2_letter}{row}"]
            
            if cell1.data_type == 'f' and 'SUM' in str(cell1.value):
                formula_cell1 = cell1
                
            if cell2.data_type == 'f' and 'SUM' in str(cell2.value):
                formula_cell2 = cell2

        if not formula_cell1 or not formula_cell2:
            logging.error("Could not find formula cells")
            return community_name, None, None

        # Extract range from formula (e.g., "=SUM(O9:O117)" -> (9, 117))
        def extract_range(formula: str) -> Tuple[Optional[int], Optional[int]]:
            match = re.search(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', formula)
            if match:
                return int(match.group(2)), int(match.group(4))
            return None, None

        start1, end1 = extract_range(formula_cell1.value)
        start2, end2 = extract_range(formula_cell2.value)

        if not all([start1, end1, start2, end2]):
            logging.error("Could not extract ranges from formulas")
            return community_name, None, None

        # Convert value sheet to DataFrame for calculations
        values_data = []
        for row in sheet_values.rows:
            values_data.append([cell.value for cell in row])
        values_df = pd.DataFrame(values_data)
        values_df.columns = [f'Column_{i}' for i in range(values_df.shape[1])]

        # Calculate total sum using formula range
        cell_value1 = get_formula_sum(values_df, column_name1, start1, end1)
        
        # Calculate sum across all country columns using the same row range
        start_col_idx = int(column_name2.split('_')[1])
        formula_result_int = get_countries_sum(values_df, start_col_idx, start2, end2)

        if cell_value1 is not None and formula_result_int is not None:
            if cell_value1 == formula_result_int:
                logging.info(f"OK {community_name}: {cell_value1} = {formula_result_int}")
            else:
                logging.warning(f"Mismatch {community_name}: {cell_value1} != {formula_result_int}")

        return community_name, cell_value1, formula_result_int

    except Exception as e:
        logging.error(f"Error processing {community_name}: {str(e)}")
        return community_name, None, None

def process_summary_sheet(file_path: str) -> Tuple[str, Optional[int], Optional[int]]:
    try:
        file_name = os.path.basename(file_path)
        community_name = extract_community_name(file_name)

        # Завантажуємо обидві версії файлу
        wb_formulas = openpyxl.load_workbook(filename=file_path, data_only=False)
        wb_values = openpyxl.load_workbook(filename=file_path, data_only=True)
        
        if 'Summary' not in wb_formulas.sheetnames:
            logging.error(f"Summary sheet not found in {community_name}")
            return community_name, None, None

        sheet_formulas = wb_formulas["Summary"]
        sheet_values = wb_values["Summary"]
        
        # Конвертуємо обидва листи в DataFrame
        data = []
        for row in sheet_formulas.rows:
            data.append([cell.value for cell in row])
        df = pd.DataFrame(data)
        df.columns = [f'Column_{i}' for i in range(df.shape[1])]

        # Конвертуємо дані значень у DataFrame
        values_data = []
        for row in sheet_values.rows:
            values_data.append([cell.value for cell in row])
        values_df = pd.DataFrame(values_data)
        values_df.columns = [f'Column_{i}' for i in range(values_df.shape[1])]

        # Шукаємо стовпці з текстом 'MESSAGING' та 'TALENT MODULE'
        messaging_col = None
        talent_module_col = None
        for col in df.columns:
            col_values = df[col].astype(str)
            if col_values.str.contains('MESSAGING', case=False, na=False).any():
                messaging_col = col
            if col_values.str.contains('TALENT MODULE', case=False, na=False).any():
                talent_module_col = col

        if messaging_col is None:
            logging.error(f"Could not find MESSAGING column in Summary sheet for {community_name}")
            return community_name, None, None

        if talent_module_col is None:
            logging.error(f"Could not find TALENT MODULE column in Summary sheet for {community_name}")
            return community_name, None, None

        # Обробка MESSAGING
        col_idx = int(messaging_col.split('_')[1])
        col_letter = get_column_letter(col_idx)
        
        messaging_sum = None
        formula_cell = None
        for row in range(1, sheet_formulas.max_row + 1):
            cell = sheet_formulas[f"{col_letter}{row}"]
            if cell.data_type == 'f' and 'SUM' in str(cell.value):
                formula_cell = cell
                break

        if formula_cell:
            match = re.search(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', formula_cell.value)
            if match:
                start_row, end_row = int(match.group(2)), int(match.group(4))
                messaging_sum = get_formula_sum(values_df, messaging_col, start_row, end_row)

        # Обробка TALENT MODULE (аналогічно до MESSAGING)
        ts_col_idx = int(talent_module_col.split('_')[1])
        ts_col_letter = get_column_letter(ts_col_idx)
        
        ts_total = None
        ts_formula_cell = None
        for row in range(1, sheet_formulas.max_row + 1):
            cell = sheet_formulas[f"{ts_col_letter}{row}"]
            if cell.data_type == 'f' and 'SUM' in str(cell.value):
                ts_formula_cell = cell
                break

        if ts_formula_cell:
            match = re.search(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', ts_formula_cell.value)
            if match:
                start_row, end_row = int(match.group(2)), int(match.group(4))
                ts_total = get_formula_sum(values_df, talent_module_col, start_row, end_row)

        return community_name, messaging_sum, ts_total

    except Exception as e:
        logging.error(f"Error processing Summary sheet for {community_name}: {str(e)}")
        return community_name, None, None

# Compare and collect results for all communities
def compare_all_communities(folder_path: str, search_text1: str, search_text2: str) -> Dict[str, Tuple[Optional[int], Optional[int]]]:
    results = {}
    logging.info("-----------------------------------------------------")
    logging.info("Verify Total = SUM per country (Messaging details sheet)")
    logging.info("-----------------------------------------------------")
    
    for file_name in os.listdir(folder_path):
        if file_name.endswith('-reporting-consumptions.xlsx') and not file_name.startswith('~$'):
            file_path = os.path.join(folder_path, file_name)
            community_name, cell_value1, formula_result_int = process_file(file_path, search_text1, search_text2)
            results[community_name] = (cell_value1, formula_result_int)
    return results

def verify_summary_totals(folder_path: str, messaging_details_results: Dict[str, Tuple[Optional[int], Optional[int]]]) -> Dict[str, Tuple[Optional[int], Optional[int]]]:
    summary_results = {}
    logging.info("-----------------------------------------------------------------")
    logging.info("Verify Total consumption (Summary sheet = Messaging details sheet)")
    logging.info("-----------------------------------------------------------------")
    
    for file_name in os.listdir(folder_path):
        if file_name.endswith('-reporting-consumptions.xlsx') and not file_name.startswith('~$'):
            file_path = os.path.join(folder_path, file_name)
            community_name, summary_value, ts_total = process_summary_sheet(file_path)
            summary_results[community_name] = (summary_value, ts_total)
            
            # Порівнюємо з результатом з Messaging details
            if community_name in messaging_details_results:
                cell_value1 = messaging_details_results[community_name][0]
                if cell_value1 is not None and summary_value is not None:
                    if cell_value1 == summary_value:
                        logging.info(f"OK {community_name}: {summary_value} = {cell_value1}")
                    else:
                        logging.warning(f"Mismatch {community_name}: {summary_value} != {cell_value1}")
                        
                else:
                    logging.error(f"Cannot compare Summary and Messaging details for {community_name} - missing values")
            else:
                logging.error(f"No Messaging details results found for {community_name}")
                
    return summary_results

def compare_with_community_sums(community_sums_df: pd.DataFrame, results: Dict[str, Tuple[Optional[int], Optional[int]]], summary_results: Dict[str, Tuple[Optional[int], Optional[int]]], community_sums_path: str):
    logging.info("------------------------------------------------------------------")
    logging.info("Compare Total SMS consumption in our reports with isendpro reports")
    logging.info("------------------------------------------------------------------")

    # Add new columns
    community_sums_df['Our report (Messaging + Talent Module)'] = None
    community_sums_df['Messaging module'] = None
    community_sums_df['Talent Module'] = None
    community_sums_df['Difference (%)'] = None
    community_sums_df['Difference (SMS)'] = None

    for community_name, (cell_value1, formula_result_int) in results.items():
        if cell_value1 is not None and formula_result_int is not None:
            try:
                row_index = community_sums_df[community_sums_df['Community name'].str.lower() == community_name.lower()].index
                if not row_index.empty:
                    total_value = community_sums_df.at[row_index[0], 'iSendPro']
                    
                    # Get ts_total value
                    summary_value, ts_total = summary_results.get(community_name, (None, None))
                    
                    # Calculate new cell_value1 as sum with ts_total
                    combined_value = cell_value1
                    if ts_total is not None:
                        combined_value += ts_total
                    
                    community_sums_df.at[row_index[0], 'Our report (Messaging + Talent Module)'] = combined_value
                    community_sums_df.at[row_index[0], 'Messaging module'] = summary_value
                    community_sums_df.at[row_index[0], 'Talent Module'] = ts_total

                    if combined_value == total_value:
                        logging.info(f"OK {community_name} ({combined_value} = {total_value})")
                    else:
                        percentage_difference = ((combined_value - total_value) / total_value) * 100
                        logging.warning(f"{community_name} ({combined_value} != {total_value}). Diff: ({percentage_difference:.1f}%)")
                        
                        community_sums_df.at[row_index[0], 'Difference (%)'] = f"=IF(C{row_index[0]+2}=0, 0, ROUND((C{row_index[0]+2}-B{row_index[0]+2})/B{row_index[0]+2}*100, 3))"
                        community_sums_df.at[row_index[0], 'Difference (SMS)'] = f"=C{row_index[0]+2}-B{row_index[0]+2}"

                else:
                    logging.error(f"Community not found: {community_name}")
            except Exception as e:
                logging.error(f"Error comparing {community_name}: {e}")
        else:
            logging.error(f"Cannot compare values for community {community_name} because cell_value1 or formula_result_int is None")

    # Зберігаємо оновлений Excel файл
    with pd.ExcelWriter(community_sums_path, engine='openpyxl') as writer:
        community_sums_df.to_excel(writer, index=False)
        
        # Отримуємо доступ до листа
        worksheet = writer.sheets['Sheet1']
        
        # Налаштовуємо ширину стовпців
        for idx, col in enumerate(community_sums_df.columns, start=0):  # start=1 для коректної нумерації стовпців Excel
            # Знаходимо максимальну довжину в стовпці
            max_length = len(str(col))  # Довжина заголовка
            
            # Перевіряємо довжину значень у стовпці
            column = community_sums_df[col]
            for value in column:
                if value is not None:
                    # Для формул враховуємо довжину результату, а не самої формули
                    if isinstance(value, str) and value.startswith('='):
                        max_length = max(max_length, 15)  # Приблизна довжина для відображення відсотків
                    else:
                        max_length = max(max_length, len(str(value)))
            
            # Додаємо більший відступ для довгих заголовків
            if col == 'Our report (Messaging + Talent Module)':
                adjusted_width = 35
                logging.info(f"Adjusted width for {col}: {adjusted_width}")
            else:
                adjusted_width = max_length + 2
                logging.info(f"Adjusted width for {col}: {adjusted_width}")
                
            # Використовуємо idx безпосередньо, оскільки тепер він починається з 1
            col_letter = get_column_letter(idx)
            worksheet.column_dimensions[col_letter].width = adjusted_width

# Main function to process all files and compare
def process_all_files_in_folder(folder_path: str, search_text1: str, search_text2: str, community_sums_path: str):
    community_sums_df = load_community_sums(community_sums_path)
    logging.info("Starting comparison...")
    
    # Перша перевірка
    results = compare_all_communities(folder_path, search_text1, search_text2)
    
    # Друга перевірка - передаємо результати з першої перевірки
    summary_results = verify_summary_totals(folder_path, results)
    
    # Порівняння з community_sums
    compare_with_community_sums(community_sums_df, results, summary_results, community_sums_path)
    logging.info("Comparison complete")

if __name__ == "__main__":
    log_file = 'process_log.txt'
    setup_logging(log_file)
    
    folder_path = 'files/'
    community_sums_path = 'result.xlsx'
    search_text1 = "SMS consumption\n(total)"
    search_text2 = "SMS consumption (distributed by Countries)"
    
    process_all_files_in_folder(folder_path, search_text1, search_text2, community_sums_path)
